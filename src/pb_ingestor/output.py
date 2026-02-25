from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .crosswalk import CrosswalkRow
from .mapper import MappedRow


def write_normalized_csv(mapped: list[MappedRow], output_path: str | Path) -> None:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    all_columns: list[str] = []
    seen: set[str] = set()
    for m in mapped:
        for k in m.row.keys():
            if k not in seen:
                seen.add(k)
                all_columns.append(k)

    with output_file.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=all_columns)
        writer.writeheader()
        for m in mapped:
            writer.writerow(m.row)


def write_manual_review_csv(mapped: list[MappedRow], output_path: str | Path) -> None:
    manual = [m.row for m in mapped if m.status != "processed"]
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    if not manual:
        output_file.write_text("Status,Status Reason\n")
        return

    cols: list[str] = []
    seen = set()
    for r in manual:
        for k in r:
            if k not in seen:
                seen.add(k)
                cols.append(k)

    with output_file.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=cols)
        writer.writeheader()
        for row in manual:
            writer.writerow(row)


def _find_header_row(ws, target_columns: list[str]):
    wanted = {c.strip().lower() for c in target_columns}
    for row_idx in range(1, min(60, ws.max_row) + 1):
        headers = [str(ws.cell(row=row_idx, column=col).value or "").strip().lower() for col in range(1, ws.max_column + 1)]
        overlap = sum(1 for h in headers if h and h in wanted)
        if overlap >= max(1, min(3, len(wanted))):
            return row_idx, headers
    return 1, [str(ws.cell(row=1, column=col).value or "").strip().lower() for col in range(1, ws.max_column + 1)]


def write_template_workbook(
    mapped: list[MappedRow],
    template_path: str | Path,
    output_workbook_path: str | Path,
    crosswalk: list[CrosswalkRow],
) -> None:
    wb = load_workbook(template_path)
    by_sheet: dict[str, list[CrosswalkRow]] = {}
    for row in crosswalk:
        by_sheet.setdefault(row.output_sheet, []).append(row)

    for sheet_name, cw_rows in by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        target_cols = [c.output_column for c in cw_rows]
        header_row, header_cells = _find_header_row(ws, target_cols)
        col_idx = {header_cells[i]: i + 1 for i in range(len(header_cells)) if header_cells[i]}

        start_row = header_row + 1
        for idx, mapped_row in enumerate(mapped, start=0):
            row_num = start_row + idx
            for cw in cw_rows:
                lc = cw.output_column.strip().lower()
                if lc not in col_idx:
                    continue
                ws.cell(row=row_num, column=col_idx[lc], value=mapped_row.row.get(cw.output_column))

    out = Path(output_workbook_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def write_qa_json(
    qa_path: str | Path,
    counters: dict[str, int],
    ingest_mode: str,
    ingest_errors: list[str],
    source_file: str,
    asset_refs: list[dict[str, str]],
) -> None:
    qa_file = Path(qa_path)
    qa_file.parent.mkdir(parents=True, exist_ok=True)

    report: dict[str, Any] = {
        "run_id": datetime.now(timezone.utc).strftime("run-%Y%m%d%H%M%S"),
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        **counters,
        "files_processed_standard": 1 if ingest_mode == "xlsx" else 0,
        "files_processed_fallback": 1 if ingest_mode == "fallback" else 0,
        "files_failed": 1 if ingest_mode == "fallback_failed" else 0,
        "rows_recovered_fallback": counters["rows_total"] if ingest_mode == "fallback" else 0,
        "rows_manual_review_fallback": counters["rows_manual_review"] if ingest_mode == "fallback" else 0,
        "summary_text": f"{counters['rows_processed']} processed / {counters['rows_incomplete']} incomplete",
        "file_results": [
            {
                "file_name": source_file,
                "status": "processed" if ingest_mode == "xlsx" else ("processed_fallback" if ingest_mode == "fallback" else "failed"),
                "rows_processed": counters["rows_processed"],
                "rows_incomplete": counters["rows_incomplete"],
                "error_message": "; ".join(ingest_errors) if ingest_errors else None,
            }
        ],
        "asset_refs": asset_refs,
    }

    qa_file.write_text(json.dumps(report, indent=2))
