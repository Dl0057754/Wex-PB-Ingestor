from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


@dataclass
class SourceRow:
    source_file: str
    source_sheet: str
    source_row_number: int
    values: dict[str, Any]
    family_context: str | None = None


@dataclass
class IngestResult:
    rows: list[SourceRow]
    mode: str
    errors: list[str]
    asset_refs: list[dict[str, str]]
    parser_stage: str = "xlsx"


def _normalize_header(value: Any, idx: int) -> str:
    if value is None:
        return f"col_{idx}"
    label = str(value).strip()
    return label if label else f"col_{idx}"


def _fill_merged_cells(ws) -> None:
    merged_ranges = list(ws.merged_cells.ranges)
    for merged in merged_ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_left = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merged))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row, col).value = top_left


def _extract_asset_refs(raw: bytes) -> list[dict[str, str]]:
    refs: list[dict[str, str]] = []
    patterns = {
        "jpg": rb"[^\x00\r\n\t ]+\.jpe?g",
        "png": rb"[^\x00\r\n\t ]+\.png",
        "pdf": rb"[^\x00\r\n\t ]+\.pdf",
        "docx": rb"[^\x00\r\n\t ]+\.docx",
    }
    for asset_type, pattern in patterns.items():
        for match in re.findall(pattern, raw, flags=re.IGNORECASE):
            try:
                ref = match.decode("latin-1")
            except Exception:
                continue
            refs.append({"asset_type": asset_type, "asset_name_or_ref": ref})
    unique = []
    seen = set()
    for r in refs:
        key = (r["asset_type"], r["asset_name_or_ref"])
        if key in seen:
            continue
        seen.add(key)
        unique.append(r)
    return unique


def _rows_from_delimited(text_lines: list[str], delimiter: str, source_file: str, stage: str) -> list[SourceRow]:
    rows: list[SourceRow] = []
    reader = csv.reader(text_lines, delimiter=delimiter)
    header = None
    for i, row in enumerate(reader, start=1):
        if not row:
            continue
        if header is None:
            header = [_normalize_header(v, idx + 1) for idx, v in enumerate(row)]
            continue
        vals = {header[idx] if idx < len(header) else f"col_{idx+1}": value for idx, value in enumerate(row)}
        if not any(str(v).strip() for v in vals.values()):
            continue
        rows.append(SourceRow(source_file=source_file, source_sheet="Recovered Sheet 1", source_row_number=i, values=vals))
    return rows


def _rows_from_fixed_width(lines: Iterable[str], source_file: str) -> list[SourceRow]:
    rows: list[SourceRow] = []
    tokenized = []
    for line in lines:
        chunks = [c for c in re.split(r"\s{2,}", line.strip()) if c]
        if len(chunks) >= 3:
            tokenized.append(chunks)
    if len(tokenized) < 2:
        return rows
    header = [_normalize_header(v, idx + 1) for idx, v in enumerate(tokenized[0])]
    for i, vals in enumerate(tokenized[1:], start=2):
        row = {header[idx] if idx < len(header) else f"col_{idx+1}": value for idx, value in enumerate(vals)}
        rows.append(SourceRow(source_file=source_file, source_sheet="Recovered Sheet 1", source_row_number=i, values=row))
    return rows


def ingest_xlsx(path: str | Path) -> IngestResult:
    rows: list[SourceRow] = []
    errors: list[str] = []
    raw = Path(path).read_bytes()
    asset_refs = _extract_asset_refs(raw)
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        return ingest_fallback(path, [f"xlsx parsing failed: {exc}"], asset_refs)

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        try:
            _fill_merged_cells(ws)
        except Exception as exc:
            errors.append(f"merge fill failed on {ws.title}: {exc}")

        header_row = None
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
            non_empty = [c for c in row if c not in (None, "")]
            if len(non_empty) >= 3:
                header_row = i
                headers = [_normalize_header(v, idx + 1) for idx, v in enumerate(row)]
                break

        if header_row is None:
            continue

        family_context = None
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            vals = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
            stripped = [str(v).strip() for v in vals.values() if v not in (None, "")]
            if not stripped:
                continue

            if len(stripped) == 1 and len(stripped[0].split()) <= 8:
                family_context = stripped[0]
                continue

            rows.append(
                SourceRow(
                    source_file=Path(path).name,
                    source_sheet=ws.title,
                    source_row_number=row_num,
                    values=vals,
                    family_context=family_context,
                )
            )

    return IngestResult(rows=rows, mode="xlsx", errors=errors, asset_refs=asset_refs, parser_stage="openxml")


def ingest_fallback(path: str | Path, pre_errors: list[str] | None = None, asset_refs: list[dict[str, str]] | None = None) -> IngestResult:
    rows: list[SourceRow] = []
    errors = list(pre_errors or [])
    raw = Path(path).read_bytes()
    if asset_refs is None:
        asset_refs = _extract_asset_refs(raw)

    text = None
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        errors.append("fallback decode failed")
        return IngestResult(rows=[], mode="fallback_failed", errors=errors, asset_refs=asset_refs, parser_stage="decode")

    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        errors.append("fallback had no non-empty lines")
        return IngestResult(rows=[], mode="fallback_failed", errors=errors, asset_refs=asset_refs, parser_stage="empty")

    attempts: list[tuple[str, list[SourceRow]]] = []
    for delim in [",", "\t", ";", "|"]:
        parsed = _rows_from_delimited(lines, delim, Path(path).name, stage=f"delim_{repr(delim)}")
        attempts.append((f"delim_{repr(delim)}", parsed))

    best_stage, best_rows = max(attempts, key=lambda item: len(item[1]))
    if len(best_rows) < 2:
        fw_rows = _rows_from_fixed_width(lines, Path(path).name)
        if len(fw_rows) > len(best_rows):
            best_rows = fw_rows
            best_stage = "fixed_width"

    if not best_rows:
        errors.append("fallback parser produced zero rows")
        return IngestResult(rows=[], mode="fallback_failed", errors=errors, asset_refs=asset_refs, parser_stage="no_rows")

    return IngestResult(rows=best_rows, mode="fallback", errors=errors, asset_refs=asset_refs, parser_stage=best_stage)
